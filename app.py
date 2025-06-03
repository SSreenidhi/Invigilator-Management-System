import streamlit as st
import sqlite3
import pandas as pd
from openpyxl import load_workbook
import re
import io
from collections import defaultdict
import time

# ---------- DB CONNECTION ----------
def get_connection():
    return sqlite3.connect("faculty.db", check_same_thread=False)

# ---------- INIT STREAMLIT ----------
st.set_page_config(page_title="Invigilation Scheduler", layout="wide")
st.title("👨‍🏫 Number of Invigilators Required")

# ---------- FILE UPLOAD ----------
teacher_upload = st.file_uploader("📤 Upload Teacher Timetable (Excel)", type=["xlsx"])
with st.expander("📘 How to Format Your Exam Timetable Excel File"):
    st.markdown("""
    ### 🗂 Required Columns in Your Excel File

    Your exam timetable must include the following **columns** exactly as listed:

    | Column Name | Description |
    |-------------|-------------|
    | **Date**    | Exam date (`DD-MM-YYYY` or `YYYY-MM-DD`) |
    | **Day**     | Day of the week (e.g., `MON`, `TUE`, `WED`, ...) |
    | **Session** | Required for **Mid exams** only. Use `FN` (Forenoon) or `AN` (Afternoon) |
    | **Subject** | Exam/course name (e.g., `Data Analysis`) |
    | **Slots**   | Comma-separated list of time slots (e.g., `I, II, III`) |

    ---
    ### ⏰ Slot Details and Timing

    Each exam slot represents a **fixed 1-hour period** in the college timetable.

    | Slot | Time               |
    |------|--------------------|
    | I    | 09:00 AM – 10:00 AM |
    | II   | 10:00 AM – 11:00 AM |
    | III  | 11:10 AM – 12:10 PM |
    | IV   | 12:10 PM – 01:10 PM |
    | V    | 12:55 PM – 01:55 PM |
    | VI   | 01:55 PM – 02:55 PM |
    | VII  | 02:55 PM – 03:45 PM |

    ---
    ### ❓ Why Slots Matter

    Slots are used to check **faculty availability**.  
    A teacher is eligible to invigilate only if they are free during **all the specified slots** of an exam.

    Example:
    > If an exam runs from `09:00 AM to 12:10 PM`, use: `Slots = I, II, III`

    ---
    ### ✅ Example Row (Excel)

    | Date       | Day | Session | Subject       | Slots       |
    |------------|-----|---------|----------------|-------------|
    | 2025-06-14 | FRI | FN      | Data Analysis | I, II, III   |


    ✅ 
    """, unsafe_allow_html=True)

exam_upload = st.file_uploader("📤 Upload Exam Timetable (Excel)", type=["xlsx"])

if not teacher_upload or not exam_upload:
    st.warning("Please upload both Teacher Timetable and Exam Timetable files.")
    st.stop()

exam_df = pd.read_excel(exam_upload)
exam_df['Date'] = pd.to_datetime(exam_df['Date'], errors='coerce')

# ---------- PARSE TEACHER TIMETABLE ONCE ----------
@st.cache_resource
def parse_teacher_timetable(file):
    wb = load_workbook(file, data_only=True)
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]
    slot_indices = {i: re.findall(r'\b[IVX]{1,4}\b', str(headers[i]))[-1] for i in range(4, 11) if re.findall(r'\b[IVX]{1,4}\b', str(headers[i]))}
    slot_to_columns = defaultdict(list)
    for col_idx, slot in slot_indices.items():
        slot_to_columns[slot].append(col_idx)
    merged_ranges = sheet.merged_cells.ranges

    def is_slot_busy(row, col_idx):
        cell = row[col_idx]
        value = cell.value
        for merged_range in merged_ranges:
            if cell.coordinate in merged_range:
                top_left = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                return top_left.value is not None and str(top_left.value).strip() != ''
        return value is not None and str(value).strip() != ''

    availability = defaultdict(lambda: defaultdict(list))
    current_teacher = None
    for row in sheet.iter_rows(min_row=2, max_col=11):
        name_cell = row[1].value
        day_cell = row[3].value
        if name_cell:
            current_teacher = str(name_cell).strip()
        if not current_teacher or not day_cell:
            continue
        day = str(day_cell).strip().upper()
        for slot, cols in slot_to_columns.items():
            if not any(is_slot_busy(row, col) for col in cols):
                availability[day][slot].append(current_teacher)
    return availability

availability_data = parse_teacher_timetable(teacher_upload)

# ---------- STREAMLIT STATE INIT ----------
if "assignment_matrix" not in st.session_state:
    st.session_state.assignment_matrix = []

# ---------- EXAM TYPE SELECTION ----------
exam_type = st.selectbox("📝 Select Exam Type", ["Semester", "Mid"])

# ---------- RESET DB ----------
conn = get_connection()
if st.button("🔄 Reset Invigilation Database"):
    cursor = conn.cursor()
    cursor.execute("UPDATE Faculty SET invigilation_count = 0")
    cursor.execute("DELETE FROM InvigilationAssignments")
    conn.commit()
    conn.close()
    st.success("✅ Database has been reset. All counts set to 0 and assignments cleared.")
    st.rerun()

cursor = conn.cursor()
assignment_rows = []
assigned_count = 0

st.markdown("### 🎯 Set Number of Invigilators for Each Exam")
exam_faculty_counts = {}

with st.form("faculty_input_form"):
    for idx, row in exam_df.iterrows():
        subject = row['Subject']
        date = row['Date']
        key = f"fac_count_{idx}"
        col1, col2 = st.columns([0.7, 0.3])
        col1.markdown(f"**{date.date() if pd.notna(date) else 'Invalid'}**")
        exam_faculty_counts[key] = col2.number_input("", min_value=1, max_value=50, value=10, key=key)
    submitted = st.form_submit_button("🔁 Generate Assignments")

if submitted:
    start = time.time()
    day_exam_group = exam_df.groupby('Day')
    for day, exams in day_exam_group:
        for idx, exam in exams.iterrows():
            slots_raw = str(exam['Slots'])
            slots = [s.strip().upper() for s in slots_raw.split(',') if s.strip()]
            subject = exam['Subject']
            date = exam['Date']

            if exam_type == "Semester":
                session = "FN"
                start_time = "10:00 AM"
                end_time = "01:00 PM"
            else:
                session = exam['Session'] if 'Session' in exam and pd.notna(exam['Session']) else "FN"
                start_time, end_time = ("10:00 AM", "12:00 PM") if session == "FN" else ("01:15 PM", "03:15 PM")

            # Get free teachers for the given slots
            try:
                common_free = list(set.intersection(*map(set, (availability_data[day][s] for s in slots if s in availability_data[day]))))
            except ValueError:
                common_free = []

            # Fetch current invigilation counts live
            cursor.execute("SELECT name, invigilation_count FROM Faculty")
            inv_count_map = {name: count for name, count in cursor.fetchall()}

            teacher_data = [{'name': t, 'count': inv_count_map.get(t, 0)} for t in common_free]
            teacher_data.sort(key=lambda x: x['count'])

            count_key = f"fac_count_{exam_df.index.get_loc(idx)}"
            required_faculty = exam_faculty_counts.get(count_key, 10)
            auto_selected = teacher_data[:required_faculty]

            for t in auto_selected:
                name = t['name']
                cursor.execute("INSERT OR IGNORE INTO Faculty (name, department, invigilation_count) VALUES (?, '', 0)", (name,))
                cursor.execute("UPDATE Faculty SET invigilation_count = invigilation_count + 1 WHERE name = ?", (name,))
                cursor.execute("SELECT invigilation_count FROM Faculty WHERE name = ?", (name,))
                inv_count = cursor.fetchone()[0]

                cursor.execute("INSERT INTO InvigilationAssignments (date, subject, day, slots, teacher) VALUES (?, ?, ?, ?, ?)",
                               (str(date), subject, day, slots_raw, name))
                assignment_rows.append({
                    "Faculty": name,
                    "Exam Day": day,
                    "Date": str(date.date()) if pd.notna(date) else "Invalid",
                    "Timings": f"{start_time} - {end_time}",
                    "Course": subject,
                    "Session": session,
                    "Invigilation Count": inv_count
                })
                assigned_count += 1

    conn.commit()

    st.write("⏱️ Time taken:", round(time.time() - start, 2), "seconds")

    if assignment_rows:
        st.success("✅ Invigilation Assignments Generated Successfully!")
        st.info(f"🎓 Assigned {assigned_count} invigilators for {len(assignment_rows)} exam slots.")
        df_out = pd.DataFrame(assignment_rows)
        st.dataframe(df_out, use_container_width=True)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_out.to_excel(writer, index=False)
        output.seek(0)

        st.download_button(
            label="📥 Download Assignment Summary (Excel)",
            data=output,
            file_name="invigilation_summary_structured.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
